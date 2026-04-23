"""
Import Router with AI-driven Schema Discovery Pipeline.
No hardcoded FIELD_TO_COLUMNS — all column mappings discovered dynamically at runtime.
"""
import asyncio
import logging
import time
import uuid
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from datetime import datetime, date
from db.manager import db_manager
from logic.lifecycle import lifecycle_manager
from logic.excel_engine import excel_engine, clean_value
from agents.column_mapper import column_mapper, SchemaDiscoveryResponse, SchemaMapping
from agents.schedule_engine import schedule_engine
from agents.status_classifier import status_classifier

logger = logging.getLogger("sidecar.routers.import_router")

router = APIRouter(prefix="/import", tags=["Import"])


# ── Request / Response models ────────────────────────────────────────────────

class SchemaConfirmRequest(BaseModel):
    """Request to confirm import with file path and approved schemas."""
    file_path: str  # Path to uploaded file
    sheet_schemas: Dict[str, List[Dict[str, Any]]]  # sheet_name -> list of SchemaMapping dicts


class ImportResult(BaseModel):
    """Result of importing a single sheet."""
    sheet_name: str
    imported: int
    skipped: int
    errors: List[str]


class ImportResponse(BaseModel):
    """Overall import response across all sheets."""
    import_id: str
    file_path: str
    results: List[ImportResult]
    total_imported: int
    total_skipped: int
    errors: List[str]


class UploadResponse(BaseModel):
    """Response to initial upload - contains discovered schema for review."""
    import_id: str
    file_path: str
    sheets: Dict[str, Any]  # sheet_name -> {headers, schema, preview_rows}
    needs_review: bool  # True if any schema has needs_review columns


def _get_value_from_row(row_data: Dict[str, Any], possible_keys: List[str], default=None):
    """Get first non-None value from row using any of the possible keys."""
    for key in possible_keys:
        val = row_data.get(key)
        if val is not None and val != "":
            return val
    return default


def _parse_months_from_string(s: str) -> Optional[int]:
    """Parse '24 MONTHS' or '2 YRS' into integer months."""
    if not s:
        return None
    import re
    # Match number followed by MONTHS, MTHS, YRS, YEARS
    match = re.search(r'(\d+)\s*(MONTHS?|MTHS?|YRS?|YEARS?)', str(s).upper())
    if match:
        num = int(match.group(1))
        unit = match.group(2)
        if 'YEAR' in unit or 'YR' in unit:
            return num * 12
        return num
    return None


async def _import_sheet(
    file_path: str,
    sheet_name: str,
    schema: List[SchemaMapping],
    conn
) -> ImportResult:
    """
    Import a single sheet using the confirmed AI-discovered schema.
    Converts Excel to CSV first for cleaner data extraction.
    """
    logger.info(f"Importing sheet '{sheet_name}' with {len(schema)} columns")
    
    # Convert Excel to CSV first
    try:
        csv_path, _, _ = excel_engine.convert_sheet_to_csv(file_path, sheet_name)
        if not csv_path:
            return ImportResult(
                sheet_name=sheet_name,
                imported=0,
                skipped=0,
                errors=[f"Failed to convert sheet '{sheet_name}' to CSV"]
            )
    except Exception as e:
        logger.error(f"Failed to convert sheet to CSV: {e}")
        return ImportResult(
            sheet_name=sheet_name,
            imported=0,
            skipped=0,
            errors=[f"CSV conversion failed: {str(e)}"]
        )
    
    imported = 0
    skipped = 0
    errors: List[str] = []
    
    # Convert schema to dict for excel_engine
    schema_dicts = [s.model_dump() for s in schema]
    
    # Extract data rows from CSV
    try:
        rows = excel_engine.extract_csv_data(csv_path, schema_dicts)
        logger.info(f"Sheet '{sheet_name}': {len(rows)} data rows extracted from CSV")
    except Exception as e:
        logger.error(f"CSV extraction failed: {e}")
        return ImportResult(
            sheet_name=sheet_name,
            imported=0,
            skipped=0,
            errors=[f"Data extraction failed: {str(e)}"]
        )
    
    if not rows:
        return ImportResult(
            sheet_name=sheet_name,
            imported=0,
            skipped=0,
            errors=["No data rows found"]
        )
    
    # Check if HRS-only asset type
    is_hrs_only = excel_engine.is_hrs_only_asset(rows)
    logger.info(f"Sheet '{sheet_name}': HRS-only = {is_hrs_only}")
    
    # Group fluid columns by fluid_type
    fluid_groups = excel_engine.group_columns_by_fluid_type(schema_dicts)
    logger.info(f"Sheet '{sheet_name}': {len(fluid_groups)} fluid types found")
    
    # Get identity/date columns for quick lookup
    ba_number_schema = next((s for s in schema if s.maps_to == "ba_number"), None)
    asset_name_schema = next((s for s in schema if s.maps_to == "asset_name"), None)
    commission_schema = next((s for s in schema if s.maps_to == "date_of_commission"), None)
    
    for row_idx, row_data in enumerate(rows):
        try:
            # ── Extract identity fields ───────────────────────────────────────
            ba_number = row_data.get("ba_number")
            logger.info(f"[IMPORT] Row {row_idx}: ba_number='{ba_number}', keys={list(row_data.keys())[:10]}...")
            
            if not ba_number:
                logger.warning(f"[IMPORT] Row {row_idx}: SKIPPED - no ba_number. Raw data: {row_data}")
                skipped += 1
                continue
            
            # Clean BA number
            ba_number = str(ba_number).strip().upper()
            
            # ── Check for duplicates ──────────────────────────────────────────
            existing = conn.execute(
                "SELECT ba_number FROM assets WHERE ba_number = ?", (ba_number,)
            ).fetchone()
            if existing:
                logger.warning(f"[IMPORT] Row {row_idx}: SKIPPED - {ba_number} already exists")
                errors.append(f"{ba_number}: already exists — skipped")
                skipped += 1
                continue
            
            # ── Extract other fields ──────────────────────────────────────────
            asset_name = row_data.get("asset_name") or ba_number
            commission_date = row_data.get("date_of_commission") or datetime.now().strftime("%Y-%m-%d")
            
            # Extract KMS (prefer kms_road, fallback to other variants)
            kms = _get_value_from_row(row_data, ["kms_road", "kms", "total_kms", "km_run"], 0.0)
            if isinstance(kms, list):  # Handle split values like [13445.0, 1042.5]
                kms = kms[0] if kms else 0.0
            
            # Extract towing KMS (second value if split)
            kms_towing = 0.0
            kms_val = row_data.get("kms_road")
            if isinstance(kms_val, list) and len(kms_val) > 1:
                kms_towing = kms_val[1]
            else:
                kms_towing = _get_value_from_row(row_data, ["kms_towing"], 0.0)
            
            # Extract HRS
            hrs = _get_value_from_row(row_data, ["hrs_run", "hrs", "hours"], 0.0)
            
            # Extract current/previous month KMS
            current_month_kms = _get_value_from_row(row_data, ["kms_current_month"], 0.0)
            previous_month_kms = _get_value_from_row(row_data, ["kms_previous_month"], 0.0)
            
            # For HRS-only assets, KMS should be 0
            if is_hrs_only:
                kms = 0.0
                kms_towing = 0.0
                current_month_kms = 0.0
                previous_month_kms = 0.0
            
            total_meterage = kms  # Lifetime total
            
            # Extract serial number
            serial_number = row_data.get("serial_number", "")
            
            # asset_group = sheet_name
            asset_group = sheet_name
            
            # asset_type from MAKE & TYPE column
            asset_type = row_data.get("asset_name", "")  # asset_name field contains MAKE & TYPE
            
            # ── INSERT asset ────────────────────────────────────────────────
            conn.execute("""
                INSERT INTO assets (
                    ba_number, name, date_of_commission,
                    serial_number, asset_group, asset_type,
                    commission_date, total_kms, kms,
                    current_month_kms, previous_month_kms, total_meterage,
                    hrs, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
            """, (
                ba_number,
                str(asset_name)[:100],  # Limit name length
                commission_date,
                str(serial_number)[:50] if serial_number else None,
                asset_group,
                str(asset_type)[:100] if asset_type else None,
                commission_date,
                float(kms) if kms else 0.0,
                float(kms) if kms else 0.0,
                float(current_month_kms) if current_month_kms else 0.0,
                float(previous_month_kms) if previous_month_kms else 0.0,
                float(total_meterage) if total_meterage else 0.0,
                float(hrs) if hrs else 0.0,
            ))
            
            # ── INSERT fluid_profiles ───────────────────────────────────────
            for fluid_type, fluid_cols in fluid_groups.items():
                profile_id = f"FLU-{ba_number}-{fluid_type}"
                
                capacity = 0.0
                top_up = 0.0
                grade = None
                last_change = None
                periodicity = None
                
                for col in fluid_cols:
                    val = row_data.get(col["maps_to"])
                    if col["maps_to"] == "fluid_capacity":
                        capacity = float(val) if val else 0.0
                    elif col["maps_to"] == "fluid_top_up":
                        top_up = float(val) if val else 0.0
                    elif col["maps_to"] == "fluid_grade":
                        grade = str(val) if val else None
                    elif col["maps_to"] == "fluid_last_change":
                        last_change = val if val else None
                    elif col["maps_to"] == "fluid_periodicity":
                        periodicity = str(val) if val else None
                
                # Only insert if we have meaningful data
                if capacity > 0 or grade:
                    conn.execute("""
                        INSERT INTO fluid_profiles (
                            profile_id, ba_number, fluid_type,
                            capacity_ltrs, top_up_10pct, grade,
                            last_change_date, periodicity
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        profile_id, ba_number, fluid_type,
                        capacity, top_up, grade, last_change, periodicity
                    ))
            
            # ── INSERT conditioning (battery) ───────────────────────────────
            battery_date = row_data.get("battery_last_change")
            battery_life = row_data.get("battery_life")
            if battery_date or battery_life:
                component_id = f"BAT-{ba_number}"
                life_months = _parse_months_from_string(str(battery_life)) if battery_life else None
                
                conn.execute("""
                    INSERT INTO components (
                        component_id, ba_number, component_type, category,
                        installed_date, life_months, status
                    ) VALUES (?, ?, 'BATTERY', 'Battery', ?, ?, 'OK')
                """, (component_id, ba_number, battery_date, life_months))
            
            # ── Schedule OH-I (with equipment type rules) ───────────────────
            lifecycle_manager.schedule_initial_overhaul(ba_number, commission_date, asset_group)
            
            # ── Seed maintenance tasks from TM history ─────────────────────
            # Check for TM-1 history
            tm1_done = row_data.get("tm1_done")
            tm1_due = row_data.get("tm1_due")
            if tm1_done and tm1_due:
                # Create completed TM-1 task
                task_id_1_done = f"TSK-{ba_number}-TM1-DONE"
                conn.execute("""
                    INSERT INTO maintenance_tasks (
                        task_id, ba_number, task_type, task_description,
                        task_interval_days, status, status_colour,
                        baseline_start_date, due_date, actual_completion_date
                    ) VALUES (?, ?, 'TM-1', 'TM-1 Service', 180, 'Completed', '#aaaaaa', ?, ?, ?)
                """, (task_id_1_done, ba_number, tm1_done, tm1_due, tm1_done))
                
                # Create scheduled next TM-1 using due date as baseline
                task_id_1_next = f"TSK-{ba_number}-TM1-NEXT"
                next_due = status_classifier.classify_task(tm1_due)
                conn.execute("""
                    INSERT INTO maintenance_tasks (
                        task_id, ba_number, task_type, task_description,
                        task_interval_days, status, status_colour,
                        baseline_start_date, due_date
                    ) VALUES (?, ?, 'TM-1', 'TM-1 Service', 180, ?, ?, ?, ?)
                """, (task_id_1_next, ba_number, 
                       next_due["status"], next_due["status_colour"],
                       tm1_done, tm1_due))
            else:
                # No history - seed fresh task
                schedule_engine.seed_initial_tasks(conn, ba_number, commission_date, is_hrs_only)
            
            # Check for TM-2 history
            tm2_done = row_data.get("tm2_done")
            tm2_due = row_data.get("tm2_due")
            if tm2_done and tm2_due:
                task_id_2_done = f"TSK-{ba_number}-TM2-DONE"
                conn.execute("""
                    INSERT INTO maintenance_tasks (
                        task_id, ba_number, task_type, task_description,
                        task_interval_days, status, status_colour,
                        baseline_start_date, due_date, actual_completion_date
                    ) VALUES (?, ?, 'TM-2', 'TM-2 Service', 365, 'Completed', '#aaaaaa', ?, ?, ?)
                """, (task_id_2_done, ba_number, tm2_done, tm2_due, tm2_done))
                
                task_id_2_next = f"TSK-{ba_number}-TM2-NEXT"
                next_due_2 = status_classifier.classify_task(tm2_due)
                conn.execute("""
                    INSERT INTO maintenance_tasks (
                        task_id, ba_number, task_type, task_description,
                        task_interval_days, status, status_colour,
                        baseline_start_date, due_date
                    ) VALUES (?, ?, 'TM-2', 'TM-2 Service', 365, ?, ?, ?, ?)
                """, (task_id_2_next, ba_number,
                       next_due_2["status"], next_due_2["status_colour"],
                       tm2_done, tm2_due))
            
            imported += 1
            logger.info(f"✅ Imported: {ba_number} from sheet '{sheet_name}'")
            
        except Exception as e:
            ba_for_error = row_data.get("ba_number", f"row-{row_idx}")
            error_msg = f"{ba_for_error}: {str(e)}"
            logger.error(f"❌ {error_msg}")
            errors.append(error_msg)
            skipped += 1
    
    conn.commit()
    logger.info(f"Sheet '{sheet_name}': imported={imported}, skipped={skipped}")
    
    return ImportResult(
        sheet_name=sheet_name,
        imported=imported,
        skipped=skipped,
        errors=errors
    )


@router.post("/upload")
async def upload_workbook(file: UploadFile = File(...)) -> UploadResponse:
    """
    STEP 1 — Upload and discover schema.
    
    1. Save uploaded file
    2. Extract headers from all sheets using openpyxl
    3. Run AI schema discovery on each sheet
    4. Return discovered schema for human review
    """
    import os
    import tempfile
    
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    import_id = f"IMP-{uuid.uuid4().hex[:8].upper()}"
    
    # Save file to temp location (cross-platform)
    temp_dir = tempfile.gettempdir()
    safe_filename = file.filename.replace(" ", "_").replace("/", "_").replace("\\", "_")
    temp_path = os.path.join(temp_dir, f"{import_id}_{safe_filename}")
    
    try:
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)
        logger.info(f"[{import_id}] Uploaded file saved to {temp_path} ({len(content)} bytes)")
    except Exception as e:
        logger.error(f"[{import_id}] Failed to save uploaded file: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
    
    try:
        # Step 1: Convert each sheet to CSV with flattened headers
        from logic.excel_engine import excel_engine
        import openpyxl
        
        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        csv_conversions = {}
        for sheet_name in sheet_names:
            csv_path, flat_headers, data_start_row = excel_engine.convert_sheet_to_csv(
                temp_path, sheet_name
            )
            if csv_path:
                csv_conversions[sheet_name] = {
                    "csv_path": csv_path,
                    "headers": flat_headers,
                    "data_start_row": data_start_row
                }
                logger.info(f"[{import_id}] CSV conversion for '{sheet_name}': {len(flat_headers)} headers, data starts at row {data_start_row}")
        
        sheets_data = {}
        needs_review = False
        
        async def process_one_sheet(sheet_name: str, sheet_info: dict) -> tuple:
            """Process a single sheet using CSV data."""
            headers = sheet_info["headers"]
            if not headers or len(headers) < 2:
                return (sheet_name, None)
            
            try:
                logger.info(f"[{import_id}] Starting schema discovery for '{sheet_name}' ({len(headers)} columns)")
                t0 = time.perf_counter()
                
                # Run AI schema discovery on flattened headers
                schema_response = await column_mapper.discover_schema(sheet_name, headers)
                
                # Get preview rows from CSV
                preview_rows = []
                try:
                    schema_dicts = [s.model_dump() for s in schema_response.column_mappings]
                    rows = excel_engine.extract_csv_data(sheet_info["csv_path"], schema_dicts)
                    preview_rows = rows[:3]
                    logger.info(f"[{import_id}] Preview rows from '{sheet_name}': {len(preview_rows)} rows")
                except Exception as e:
                    logger.warning(f"Could not extract preview for {sheet_name}: {e}")
                
                elapsed = time.perf_counter() - t0
                logger.info(f"[{import_id}] Sheet '{sheet_name}' complete in {elapsed:.1f}s")
                
                return (sheet_name, {
                    "headers": headers,
                    "csv_path": sheet_info["csv_path"],
                    "schema": [s.model_dump() for s in schema_response.column_mappings],
                    "preview_rows": preview_rows,
                    "column_count": len(headers),
                    "data_start_row": sheet_info["data_start_row"],
                    "identity_columns": [s.model_dump() for s in schema_response.get_identity_columns()],
                    "fluid_columns": [s.model_dump() for s in schema_response.get_fluid_columns()],
                    "needs_review": any(s.needs_review for s in schema_response.column_mappings),
                })
            except Exception as e:
                logger.error(f"[{import_id}] Failed to process sheet '{sheet_name}': {e}")
                return (sheet_name, None)
        
        # Process all sheets in PARALLEL
        sheet_tasks = [process_one_sheet(name, info) for name, info in csv_conversions.items()]
        results = await asyncio.gather(*sheet_tasks)
        
        # Collect results
        for sheet_name, sheet_data in results:
            if sheet_data:
                sheets_data[sheet_name] = sheet_data
                if sheet_data.get("needs_review"):
                    needs_review = True
        
        logger.info(f"[{import_id}] Discovered schema for {len(sheets_data)} sheets")
        
        return UploadResponse(
            import_id=import_id,
            file_path=temp_path,
            sheets=sheets_data,
            needs_review=needs_review
        )
        
    except Exception as e:
        logger.error(f"[{import_id}] Schema discovery failed: {e}")
        raise HTTPException(status_code=500, detail=f"Schema discovery failed: {str(e)}")


@router.post("/{import_id}/confirm", response_model=ImportResponse)
async def confirm_import(import_id: str, request: SchemaConfirmRequest) -> ImportResponse:
    """
    STEP 2 — Confirm and import with AI-discovered and user-approved schema.
    
    Backend re-reads the workbook file using confirmed schema mappings.
    """
    logger.info(f"[{import_id}] Confirming import for file: {request.file_path}")
    
    conn = db_manager.connect()
    results = []
    total_imported = 0
    total_skipped = 0
    all_errors = []
    
    for sheet_name, schema_list in request.sheet_schemas.items():
        # Convert dict list to SchemaMapping objects
        schema = [SchemaMapping(**s) for s in schema_list]
        
        # Import this sheet by re-reading the file
        result = await _import_sheet(request.file_path, sheet_name, schema, conn)
        results.append(result)
        total_imported += result.imported
        total_skipped += result.skipped
        all_errors.extend(result.errors)
    
    # Log audit
    db_manager.log_agent_action(
        agent_id="AGT-04",
        action_type="import_confirm",
        input_data={"import_id": import_id, "sheets": list(request.sheet_schemas.keys())},
        output_data={
            "total_imported": total_imported,
            "total_skipped": total_skipped,
            "sheet_count": len(results)
        },
    )
    
    logger.info(f"[{import_id}] Import complete: {total_imported} imported, {total_skipped} skipped")
    
    return ImportResponse(
        import_id=import_id,
        file_path=request.file_path,
        results=results,
        total_imported=total_imported,
        total_skipped=total_skipped,
        errors=all_errors
    )
