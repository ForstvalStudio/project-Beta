"""
Excel Engine with openpyxl for dynamic schema discovery.
Handles merged multi-row headers from production workbooks.
"""
import re
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

logger = logging.getLogger("sidecar.excel_engine")

def clean_value(v: Any) -> Any:
    """
    Clean a cell value according to the rules:
    - datetime → YYYY-MM-DD string
    - int/float > 1000 → treat as Excel serial, convert to date
    - string DD/MM/YY or DD/MM/YYYY → parse as date
    - "__", "___", "-", "_", "BOH", "None", "" → None
    - "28.000LTR" → extract float 28.0
    - "13445/1042.5" → split on "/", return [13445.0, 1042.5]
    - strip newlines and whitespace
    """
    if v is None:
        return None
    
    # Already a datetime
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    
    # Numeric Excel serial date (> 1000)
    if isinstance(v, (int, float)):
        if v > 1000 and v < 50000:  # Excel serial date range
            try:
                dt = datetime(1899, 12, 30) + timedelta(days=int(v))
                return dt.strftime("%Y-%m-%d")
            except:
                return float(v)
        return float(v)
    
    # String processing
    if isinstance(v, str):
        s = v.strip()
        
        # Empty/null indicators
        if s in ["__", "___", "-", "_", "BOH", "None", "", " "]:
            return None
        
        # Strip newlines
        s = s.replace("\n", " ").replace("\r", "").strip()
        
        # Date patterns DD/MM/YY or DD/MM/YYYY
        date_match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
        if date_match:
            day, month, year = date_match.groups()
            if len(year) == 2:
                year_int = int(year)
                if year_int >= 50:
                    year = "19" + year
                else:
                    year = "20" + year
            try:
                dt = datetime(int(year), int(month), int(day))
                return dt.strftime("%Y-%m-%d")
            except:
                pass
        
        # Extract float from strings like "28.000LTR" or "28 LTR"
        float_match = re.search(r'(\d+\.?\d*)', s.replace(",", ""))
        if float_match:
            # Check if it's a split value like "13445/1042.5"
            if "/" in s:
                parts = s.split("/")
                try:
                    return [float(re.search(r'(\d+\.?\d*)', p.strip().replace(",", "")).group(1)) 
                            for p in parts if re.search(r'(\d+\.?\d*)', p.strip().replace(",", ""))]
                except:
                    pass
            try:
                return float(float_match.group(1))
            except:
                pass
        
        return s if s else None
    
    return v


class ExcelEngine:
    """
    Excel engine using openpyxl for dynamic header extraction.
    Handles merged multi-row headers from production workbooks.
    """
    
    def __init__(self):
        logger.info("ExcelEngine initialized (openpyxl-based)")

    def extract_headers_from_sheet(self, sheet) -> List[str]:
        """
        Extract concatenated headers from the first 5 rows of a sheet.
        Handles merged cells by concatenating non-None values column by column.
        
        Example:
            Row 1: "ENG OIL", None, None, None, None
            Row 2: "CAPACITY LTRS/KG", "ADDL 10% TOP UP", ...
        Becomes: ["ENG OIL CAPACITY LTRS/KG", "ENG OIL ADDL 10% TOP UP", ...]
        """
        import openpyxl
        
        headers = []
        max_col = sheet.max_column
        
        # For each column, collect all non-None values from first 5 rows
        for col_idx in range(1, max_col + 1):
            parts = []
            for row_idx in range(1, min(6, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Handle merged cells
                if cell.value is not None:
                    val = str(cell.value).strip()
                    val = val.replace("\n", " ").replace("\r", "").strip()
                    if val and val not in parts:  # Avoid duplicates
                        parts.append(val)
            
            # Concatenate all parts with space separator
            header = " ".join(parts) if parts else f"COLUMN_{col_idx}"
            headers.append(header)
        
        logger.info(f"Extracted {len(headers)} headers from sheet: {headers[:5]}...")
        return headers

    def read_workbook_sheets(self, file_path: str) -> Dict[str, List[str]]:
        """
        Read all sheets from a workbook and return {sheet_name: headers}.
        """
        import openpyxl
        
        logger.info(f"Reading workbook: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            result = {}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = self.extract_headers_from_sheet(sheet)
                result[sheet_name] = headers
                logger.info(f"Sheet '{sheet_name}': {len(headers)} columns")
            
            wb.close()
            return result
        except Exception as e:
            logger.error(f"Failed to read workbook {file_path}: {e}")
            raise

    def convert_sheet_to_csv(
        self,
        file_path: str,
        sheet_name: str,
        output_path: Optional[str] = None
    ) -> Tuple[str, List[str], int]:
        """
        Convert Excel sheet to clean CSV with flattened headers.
        
        Handles multi-row headers by combining them into single column names.
        Auto-detects where data rows actually start (first row with numeric SER NO).
        
        Returns:
            (csv_path, flattened_headers, data_start_row)
        """
        import openpyxl
        import csv
        
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        sheet = wb[sheet_name]
        
        # Read all rows into memory to analyze structure
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            wb.close()
            return "", [], 0
        
        # Find data start row: first row where first cell is a number (SER NO)
        data_start_row = 0
        for i, row in enumerate(all_rows):
            first_cell = row[0] if row else None
            # Check if first cell is a number (asset serial number)
            if first_cell is not None:
                try:
                    float(first_cell)  # If it's numeric, this is data
                    data_start_row = i
                    break
                except (ValueError, TypeError):
                    pass
        
        if data_start_row == 0:
            # Couldn't find data start, assume standard 2 header rows
            data_start_row = 2
        
        # Build flattened headers from all rows above data
        header_rows = all_rows[:data_start_row]
        num_cols = max(len(r) for r in all_rows) if all_rows else 0
        
        flattened_headers = []
        for col_idx in range(num_cols):
            parts = []
            for row in header_rows:
                cell_val = row[col_idx] if col_idx < len(row) else None
                if cell_val is not None and str(cell_val).strip():
                    parts.append(str(cell_val).strip())
            # Combine header parts
            if parts:
                header = " ".join(parts)
            else:
                header = f"Column_{col_idx + 1}"
            flattened_headers.append(header)
        
        # Generate output path if not provided
        if output_path is None:
            import tempfile
            import os
            temp_dir = tempfile.gettempdir()
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            safe_sheet = sheet_name.replace(" ", "_").replace("&", "and")
            output_path = os.path.join(temp_dir, f"{base_name}_{safe_sheet}.csv")
        
        # Write CSV with flattened headers and data rows
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(flattened_headers)
            for row in all_rows[data_start_row:]:
                # Pad row to match header count
                padded_row = list(row) + [''] * (num_cols - len(row))
                writer.writerow(padded_row[:num_cols])
        
        wb.close()
        logger.info(f"[CSV] Converted '{sheet_name}' to {output_path}: {len(flattened_headers)} cols, {len(all_rows) - data_start_row} data rows")
        return output_path, flattened_headers, data_start_row

    def extract_csv_data(
        self,
        csv_path: str,
        schema: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Extract data rows from CSV using the AI-discovered schema.
        
        Args:
            csv_path: Path to CSV file
            schema: List of {col_index, header, category, maps_to, fluid_type}
        
        Returns:
            List of asset data dicts
        """
        import csv
        
        logger.info(f"[CSV] Extracting data from {csv_path} with {len(schema)} mapped columns")
        
        # Build column index mapping
        col_mapping = {s["col_index"]: s for s in schema if s.get("maps_to")}  # 0-based, only mapped columns
        
        # Find ba_number column for row validation
        ba_number_col = None
        for s in schema:
            if s.get("maps_to") == "ba_number":
                ba_number_col = s["col_index"]
                break
        
        rows = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)  # Skip header row
            
            for row_idx, row in enumerate(reader, start=2):
                # Skip empty rows
                if not row or not any(cell.strip() for cell in row):
                    continue
                
                # Validate ba_number exists and is not a header repeat
                if ba_number_col is not None:
                    if ba_number_col >= len(row):
                        continue
                    ba_val = clean_value(row[ba_number_col])
                    if ba_val is None or str(ba_val).strip() == "":
                        continue
                    ba_str = str(ba_val).upper()
                    if "BA NO" in ba_str or "SER" in ba_str or "VEHICLE" in ba_str:
                        continue
                
                # Build row data
                row_data = {"_row_idx": row_idx}
                for col_idx, col_schema in col_mapping.items():
                    if col_idx < len(row):
                        raw_val = row[col_idx]
                        cleaned = clean_value(raw_val)
                        row_data[col_schema["maps_to"]] = cleaned
                
                rows.append(row_data)
        
        logger.info(f"[CSV] Extracted {len(rows)} valid data rows")
        return rows

    def extract_sheet_data(
        self, 
        file_path: str, 
        sheet_name: str,
        schema: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Extract data rows from a sheet using the AI-discovered schema.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read
            schema: List of {col_index, header, category, maps_to, fluid_type} from AI
        
        Returns:
            List of asset data dicts
        """
        import openpyxl
        
        logger.info(f"Extracting data from sheet '{sheet_name}' with schema of {len(schema)} columns")
        
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet = wb[sheet_name]
        
        # Build column index mapping
        col_mapping = {s["col_index"]: s for s in schema}  # 0-based index
        
        # Find ba_number column index for row filtering
        ba_number_col = None
        ser_no_col = None  # First column (usually SER NO)
        for s in schema:
            if s.get("maps_to") == "ba_number":
                ba_number_col = s["col_index"]  # 0-based
            # Track first column as SER NO for blank row detection
            if s["col_index"] == 0:
                ser_no_col = 0
        
        rows = []
        header_row_count = 5  # First 5 rows are headers
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_count + 1, values_only=True), 
                                       start=header_row_count + 1):
            
            # Skip completely empty rows
            if not any(cell is not None for cell in row):
                continue
            
            # Skip if first cell (SER NO) is None (blank spacer row)
            if ser_no_col is not None and (len(row) <= ser_no_col or row[ser_no_col] is None):
                continue
            
            # Check if ba_number is valid
            if ba_number_col is not None:
                if len(row) <= ba_number_col:
                    continue
                ba_val = clean_value(row[ba_number_col])
                if ba_val is None or str(ba_val).strip() == "":
                    continue
                # Skip header repeats
                ba_str = str(ba_val).upper()
                if "BA NO" in ba_str or "SER" in ba_str or "VEHICLE" in ba_str:
                    continue
            
            # Build row data using schema
            row_data = {
                "_sheet_name": sheet_name,
                "_row_idx": row_idx,
            }
            
            for col_idx, col_schema in col_mapping.items():
                if len(row) > col_idx:
                    raw_val = row[col_idx]
                    cleaned = clean_value(raw_val)
                    row_data[col_schema["maps_to"]] = cleaned
                    # Also store the original header for debugging
                    row_data[f"_header_{col_schema['maps_to']}"] = col_schema["header"]
            
            rows.append(row_data)
        
        wb.close()
        logger.info(f"Extracted {len(rows)} valid data rows from sheet '{sheet_name}'")
        return rows

    def is_hrs_only_asset(self, rows: List[Dict[str, Any]]) -> bool:
        """
        Determine if this is an HRS-only asset (no KMs column or KMs is always "-").
        New equipment types: Gen set, JCB, Dozer, SSL
        """
        if not rows:
            return False
        
        kms_values = []
        for row in rows:
            # Check various possible KMS field names
            for key in ["kms_road", "kms", "total_kms", "km_run"]:
                val = row.get(key)
                if val is not None:
                    kms_values.append(val)
        
        # If no KMS values found or all are 0/None/-, it's HRS-only
        if not kms_values:
            return True
        
        non_zero_kms = [v for v in kms_values if v not in [None, 0, "-", "", 0.0]]
        return len(non_zero_kms) == 0

    def group_columns_by_fluid_type(self, schema: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Group FLUID category columns by their fluid_type.
        Returns {fluid_type: [col_schemas]}.
        """
        fluid_groups = {}
        
        for col in schema:
            if col.get("category") == "FLUID":
                fluid_type = col.get("fluid_type", "OTHER")
                if fluid_type not in fluid_groups:
                    fluid_groups[fluid_type] = []
                fluid_groups[fluid_type].append(col)
        
        return fluid_groups


# Global instance for shared use
excel_engine = ExcelEngine()
